"""
Load_Targets_From_Template.py
Loads a completed targets template into Input.Targets.

Tenant ID and FY year are read from the filename — the file must be named
T{tenant_id}_{fy}.xlsx  e.g.  T11_2025-26.xlsx

Usage:
    py Scripts/Load_Targets_From_Template.py ^
        --server <fabric-sql-endpoint> ^
        --database WH_Dentally ^
        --input Targets/T11_2025-26.xlsx ^
        [--auth interactive]  [--dry-run]

WHAT IS DELETED:
    All existing Input.Targets rows for this Tenant_ID where
    Period_Type = 'annual' AND Period_Value = the FY from the filename.
    Other Period_Types are untouched.

AFTER LOADING:
    1. Run Gold.usp_Load_Fact_Targets
    2. Run Gold.usp_Load_Fact_Effective_Targets
    3. Refresh Power BI dataset
"""

import argparse
import re
import struct
import sys
from datetime import datetime, timezone
from pathlib import Path

import pyodbc
import openpyxl


# ── Auth ─────────────────────────────────────────────────────────────────────

def _token_bytes(token_str):
    b = token_str.encode("utf-16-le")
    return struct.pack(f"<I{len(b)}s", len(b), b)


def connect(server, database, username=None, auth="default"):
    from azure.identity import DefaultAzureCredential, InteractiveBrowserCredential
    cred = (InteractiveBrowserCredential(login_hint=username)
            if auth == "interactive" else DefaultAzureCredential())
    token = cred.get_token("https://database.windows.net/.default")
    conn = pyodbc.connect(
        "Driver={ODBC Driver 18 for SQL Server};"
        f"Server={server},1433;"
        f"Database={database};"
        "Encrypt=yes;",
        attrs_before={1256: _token_bytes(token.token)},
    )
    conn.autocommit = True
    return conn


# ── Filename parser ───────────────────────────────────────────────────────────

FILENAME_RE = re.compile(r"^T(\d+)_(\d{4}-\d{2})\.xlsx$", re.IGNORECASE)


def parse_filename(path):
    m = FILENAME_RE.match(Path(path).name)
    if not m:
        sys.exit(
            f"Filename '{Path(path).name}' does not match expected pattern T{{tenant}}_{{fy}}.xlsx\n"
            f"Example: T11_2025-26.xlsx"
        )
    return int(m.group(1)), m.group(2)


# ── Row parser ────────────────────────────────────────────────────────────────

def parse_num(val, label, row_num):
    if val is None or str(val).strip() == "":
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        print(f"  Warning row {row_num}: non-numeric {label} '{val}' — skipping row")
        return "INVALID"


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--server",   required=True)
    p.add_argument("--database", required=True)
    p.add_argument("--input",    required=True, help="Path to completed template, e.g. Targets/T11_2025-26.xlsx")
    p.add_argument("--username", default=None,
                   help="Azure AD login hint (used with --auth interactive)")
    p.add_argument("--auth",     default="default", choices=["default", "interactive"])
    p.add_argument("--dry-run",  action="store_true", help="Validate without writing")
    args = p.parse_args()

    # ── Parse filename ────────────────────────────────────────────────────────
    tenant_id, fy = parse_filename(args.input)
    period_value = "FY " + fy   # filename '2025-26' → DB Period_Value 'FY 2025-26'
    print(f"File:   {args.input}")
    print(f"Tenant: {tenant_id}   FY: {fy}  Period_Value: {period_value}")

    # ── Read template ─────────────────────────────────────────────────────────
    try:
        wb = openpyxl.load_workbook(args.input, read_only=True, data_only=True)
    except FileNotFoundError:
        sys.exit(f"File not found: {args.input}")

    if "Targets" not in wb.sheetnames:
        sys.exit("Sheet 'Targets' not found.")

    ws = wb["Targets"]
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        sys.exit("Targets sheet is empty.")

    headers = [str(h).strip() for h in all_rows[0]]
    required = {"Level", "Site_ID", "Practitioner_ID", "Metric_Key", "Target_Value", "Variance_Band"}
    missing = required - set(headers)
    if missing:
        sys.exit(f"Missing columns: {missing}")

    col = {h: i for i, h in enumerate(headers)}

    # ── Parse rows ────────────────────────────────────────────────────────────
    records = []
    skipped_blank   = 0
    skipped_invalid = 0

    for row_num, row in enumerate(all_rows[1:], 2):
        raw_target = row[col["Target_Value"]]

        if raw_target is None or str(raw_target).strip() == "":
            skipped_blank += 1
            continue

        target = parse_num(raw_target, "Target_Value", row_num)
        if target == "INVALID":
            skipped_invalid += 1
            continue

        variance = parse_num(row[col["Variance_Band"]], "Variance_Band", row_num)
        if variance == "INVALID":
            variance = None

        level   = str(row[col["Level"]]).strip()
        mk      = str(row[col["Metric_Key"]]).strip()

        raw_site = row[col["Site_ID"]]
        raw_prac = row[col["Practitioner_ID"]]

        site_id = str(raw_site).strip() if raw_site and str(raw_site).strip() else None
        prac_id = int(raw_prac) if raw_prac and str(raw_prac).strip() else None

        # Enforce hierarchy — practitioners are not site-bound
        if level == "Practice":
            site_id = None
            prac_id = None
        elif level == "Site":
            prac_id = None
        elif level == "Practitioner":
            site_id = None

        records.append((tenant_id, site_id, prac_id, mk, "annual", period_value, target, variance))

    print(f"\nRows with targets: {len(records)}")
    print(f"Skipped (blank):   {skipped_blank}")
    print(f"Skipped (invalid): {skipped_invalid}")

    if not records:
        print("Nothing to load.")
        return

    # ── Summary ───────────────────────────────────────────────────────────────
    by_level = {}
    for r in records:
        lv = "Practice" if r[1] is None else ("Practitioner" if r[2] else "Site")
        by_level[lv] = by_level.get(lv, 0) + 1
    for lv, cnt in sorted(by_level.items()):
        print(f"  {lv}: {cnt} rows")

    if args.dry_run:
        print("\nDry run — no changes written.")
        return

    # ── Write ─────────────────────────────────────────────────────────────────
    print(f"\nConnecting to {args.server} / {args.database} ...")
    try:
        conn = connect(args.server, args.database, args.username, args.auth)
    except Exception as e:
        sys.exit(f"Connection failed: {e}")

    cur = conn.cursor()

    cur.execute(
        "DELETE FROM Input.Targets "
        "WHERE Tenant_ID = ? AND Period_Type = 'annual' AND Period_Value = ?",
        [tenant_id, period_value]
    )

    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    cur.fast_executemany = True
    cur.executemany(
        "INSERT INTO Input.Targets "
        "(Tenant_ID, Site_ID, Practitioner_ID, Metric, Period_Type, Period_Value, "
        " Target_Value, Variance, DW_Created_At, DW_Updated_At) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        [(r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7], now, now) for r in records],
    )

    print(f"Inserted {len(records)} rows into Input.Targets.")
    print("\nNext steps:")
    print("  1. Run Gold.usp_Load_Fact_Targets")
    print("  2. Run Gold.usp_Load_Fact_Effective_Targets")
    print("  3. Refresh Power BI dataset")

    conn.close()


if __name__ == "__main__":
    main()
