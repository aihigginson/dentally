"""
Generate_Mapping_Template.py
Generates an Excel mapping template for a given entity and tenant,
pre-populated with the distinct source values from the Fabric warehouse.
The Standard column contains a dropdown validated against Config.

Entities:  treatment_category | acquisition_source | cancellation_reason | payment_plan

Usage:
    py Scripts/Generate_Mapping_Template.py ^
        --server <fabric-sql-endpoint> ^
        --database WH_Dentally ^
        --tenant 11 ^
        --entity treatment_category ^
        [--auth interactive]

Output: Mappings/T{tenant_id}_{entity}.xlsx
Fill in the Standard column then run Load_Mapping_From_Template.py.
Existing mappings are pre-filled.
"""

import argparse
import struct
import sys
from datetime import date
from pathlib import Path

import pyodbc
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Entity configuration ──────────────────────────────────────────────────────

ENTITY_CONFIG = {
    'treatment_category': {
        'display_name':    'Treatment Category',
        'input_table':     'Input.Treatment_Category_Map',
        'config_table':    'Config.Treatment_Category_Standard',
        'source_col':      'Source_Treatment_Category',
        'standard_col':    'Standard_Treatment_Category',
        'source_query':    (
            "SELECT DISTINCT Treatment_Category_Name "
            "FROM Gold.Dim_Treatments "
            "WHERE Tenant_ID = ? AND Treatment_Category_Name IS NOT NULL "
            "ORDER BY Treatment_Category_Name"
        ),
    },
    'acquisition_source': {
        'display_name':    'Acquisition Source',
        'input_table':     'Input.Acquisition_Source_Map',
        'config_table':    'Config.Acquisition_Source_Standard',
        'source_col':      'Source_Acquisition_Source',
        'standard_col':    'Standard_Acquisition_Source',
        'source_query':    (
            "SELECT DISTINCT Name "
            "FROM Gold.Dim_Acquisition_Sources "
            "WHERE Tenant_ID = ? AND Name IS NOT NULL "
            "ORDER BY Name"
        ),
    },
    'cancellation_reason': {
        'display_name':    'Cancellation Reason',
        'input_table':     'Input.Cancellation_Reason_Map',
        'config_table':    'Config.Cancellation_Reason_Standard',
        'source_col':      'Source_Cancellation_Reason',
        'standard_col':    'Standard_Cancellation_Reason',
        'source_query':    (
            "SELECT DISTINCT Reason "
            "FROM Gold.Dim_Cancellation_Reasons "
            "WHERE Tenant_ID = ? AND Reason IS NOT NULL "
            "ORDER BY Reason"
        ),
    },
    'payment_plan': {
        'display_name':    'Payment Plan',
        'input_table':     'Input.Payment_Plan_Map',
        'config_table':    'Config.Payment_Plan_Standard',
        'source_col':      'Source_Payment_Plan',
        'standard_col':    'Standard_Payment_Plan',
        'source_query':    (
            "SELECT DISTINCT Payment_Plan_Name "
            "FROM Gold.Dim_Payment_Plans "
            "WHERE Tenant_ID = ? AND Payment_Plan_Name IS NOT NULL "
            "ORDER BY Payment_Plan_Name"
        ),
    },
}

# ── Auth ──────────────────────────────────────────────────────────────────────

def _token_bytes(token_str):
    b = token_str.encode('utf-16-le')
    return struct.pack(f'<I{len(b)}s', len(b), b)


def connect(server, database, username=None, auth='default'):
    from azure.identity import DefaultAzureCredential, InteractiveBrowserCredential
    cred = (InteractiveBrowserCredential(login_hint=username)
            if auth == 'interactive' else DefaultAzureCredential())
    token = cred.get_token('https://database.windows.net/.default')
    return pyodbc.connect(
        'Driver={ODBC Driver 18 for SQL Server};'
        f'Server={server},1433;'
        f'Database={database};'
        'Encrypt=yes;',
        attrs_before={1256: _token_bytes(token.token)},
    )

# ── Styles ────────────────────────────────────────────────────────────────────

NAVY   = '1A527A'
TEAL   = '5EC4C0'
LIGHT  = 'EBF3F9'

def _header_font():   return Font(bold=True, color='FFFFFF', size=11)
def _header_fill(c):  return PatternFill('solid', fgColor=c)
def _body_align():    return Alignment(vertical='center')


def _col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument('--server',   required=True)
    p.add_argument('--database', required=True)
    p.add_argument('--tenant',   required=True, type=int)
    p.add_argument('--entity',   required=True, choices=list(ENTITY_CONFIG))
    p.add_argument('--username', default=None)
    p.add_argument('--auth',     default='default', choices=['default', 'interactive'])
    args = p.parse_args()

    cfg = ENTITY_CONFIG[args.entity]

    print(f'Connecting to {args.server} / {args.database} ...')
    try:
        conn = connect(args.server, args.database, args.username, args.auth)
    except Exception as e:
        print(f'Connection failed: {e}', file=sys.stderr)
        sys.exit(1)

    cur = conn.cursor()

    # ── Fetch allowed standard values ─────────────────────────────────────────
    standard_col = cfg['standard_col']
    cur.execute(
        f"SELECT [{standard_col}] "
        f"FROM {cfg['config_table']} "
        f"WHERE Is_Active = 1 "
        f"ORDER BY Display_Order"
    )
    standard_values = [r[0] for r in cur.fetchall()]
    if not standard_values:
        print(f"No active standard values in {cfg['config_table']} — run the Data.sql seed first.",
              file=sys.stderr)
        sys.exit(1)

    # ── Fetch source values for this tenant ───────────────────────────────────
    cur.execute(cfg['source_query'], [args.tenant])
    source_values = [r[0] for r in cur.fetchall()]
    if not source_values:
        print(f'No source values found for tenant {args.tenant}.', file=sys.stderr)
        sys.exit(1)

    # ── Fetch existing mappings to pre-fill ───────────────────────────────────
    source_col = cfg['source_col']
    cur.execute(
        f"SELECT [{source_col}], [{standard_col}] "
        f"FROM {cfg['input_table']} "
        f"WHERE Tenant_ID = ?",
        [args.tenant],
    )
    existing = {r[0]: r[1] for r in cur.fetchall()}
    conn.close()

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = cfg['display_name']

    # Header row
    ws.append([f'Source {cfg["display_name"]}', f'Standard {cfg["display_name"]}'])
    for col, fill_hex in enumerate([NAVY, TEAL], start=1):
        cell = ws.cell(1, col)
        cell.font = _header_font()
        cell.fill = _header_fill(fill_hex)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[1].height = 22
    _col_width(ws, 1, 45)
    _col_width(ws, 2, 35)
    ws.freeze_panes = 'A2'

    # Data rows
    for i, source in enumerate(source_values, start=2):
        ws.cell(i, 1, source)
        ws.cell(i, 1).alignment = _body_align()
        ws.cell(i, 2, existing.get(source, ''))
        ws.cell(i, 2).alignment = _body_align()

    # Dropdown validation on Standard column
    quoted = ','.join(f'"{v}"' for v in standard_values)
    dv = DataValidation(
        type='list',
        formula1=f'"{quoted}"' if len(quoted) <= 255 else None,
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle='Invalid value',
        error=f'Choose a value from the standard list.',
    )
    # If the quoted list exceeds 255 chars, write standards to a hidden sheet
    if len(quoted) > 255:
        ref_ws = wb.create_sheet('_Standards')
        ref_ws.sheet_state = 'hidden'
        for j, v in enumerate(standard_values, start=1):
            ref_ws.cell(j, 1, v)
        ref_range = f'_Standards!$A$1:$A${len(standard_values)}'
        dv = DataValidation(
            type='list', formula1=ref_range, allow_blank=True,
            showDropDown=False, showErrorMessage=True,
            errorTitle='Invalid value', error='Choose a value from the standard list.',
        )

    ws.add_data_validation(dv)
    dv.add(f'B2:B{max(len(source_values) + 1, 100)}')

    # Lock source column
    ws.protection.sheet = False

    # ── Save ──────────────────────────────────────────────────────────────────
    out_dir = Path('Mappings')
    out_dir.mkdir(exist_ok=True)
    out_path = out_dir / f'T{args.tenant}_{args.entity}.xlsx'
    wb.save(out_path)
    print(f'Saved {out_path}  ({len(source_values)} source values, {len(existing)} pre-filled)')


if __name__ == '__main__':
    main()
