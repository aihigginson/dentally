"""
Generate_Targets_Template.py
Generates one Excel targets template per FY year for a given tenant,
pre-populated with sites and practitioners from the Fabric warehouse.

Output files are saved to the Targets/ folder at the repo root.
Filename format: T{tenant_id}_{fy}.xlsx  e.g. T11_2025-26.xlsx
The loader reads tenant and FY back from the filename.

Usage:
    py Scripts/Generate_Targets_Template.py ^
        --server <fabric-sql-endpoint> ^
        --database WH_Dentally ^
        --tenant 11 ^
        [--auth interactive]  [--years 3]

Fill in Target_Value (and optionally Variance_Band) then run
Load_Targets_From_Template.py.  Rows left blank are skipped.
If annual targets already exist in Input.Targets they are pre-filled.
"""

import argparse
import re
import struct
import sys
from pathlib import Path


def _norm_fy(pv):
    """Normalise any Period_Value variant to 'FY YYYY-YY'."""
    if pv is None:
        return None
    s = str(pv).strip()
    if re.match(r"^FY \d{4}-\d{2}$", s):          # already correct
        return s
    m = re.match(r"^(\d{4})-(\d{2})$", s)          # '2025-26'
    if m:
        return f"FY {m.group(1)}-{m.group(2)}"
    m = re.match(r"^FY(\d{4})[/-](\d{2})$", s)     # 'FY2025/26' or 'FY2025-26'
    if m:
        return f"FY {m.group(1)}-{m.group(2)}"
    return s


def _prior_fy_pv(fy):
    """'2025-26' -> 'FY 2024-25' (prior financial year as a Period_Value)."""
    m = re.match(r"^(\d{4})-(\d{2})$", fy)
    if not m:
        return None
    return f"FY {int(m.group(1)) - 1}-{int(m.group(2)) - 1:02d}"

import pyodbc
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ── Auth ─────────────────────────────────────────────────────────────────────

def _token_bytes(token_str):
    b = token_str.encode("utf-16-le")
    return struct.pack(f"<I{len(b)}s", len(b), b)


def connect(server, database, username=None, auth="default"):
    from azure.identity import DefaultAzureCredential, InteractiveBrowserCredential
    cred = (InteractiveBrowserCredential(login_hint=username)
            if auth == "interactive" else DefaultAzureCredential())
    token = cred.get_token("https://database.windows.net/.default")
    return pyodbc.connect(
        "Driver={ODBC Driver 18 for SQL Server};"
        f"Server={server},1433;"
        f"Database={database};"
        "Encrypt=yes;",
        attrs_before={1256: _token_bytes(token.token)},
    )


# ── Styling ───────────────────────────────────────────────────────────────────

LEVEL_FILLS = {
    "Practice":     "DCE8F5",
    "Site":         "FFF3CD",
    "Practitioner": "D5F0D5",
}
HEADER_FILL = PatternFill("solid", fgColor="1F497D")
HEADER_FONT = Font(color="FFFFFF", bold=True)

HEADERS = [
    "Practice_Name", "Level",
    "Site_ID", "Site_Name", "Practitioner_ID", "Practitioner_Name",
    "Metric_Key", "Metric_Name", "Section", "Format_Type",
    "Prior_Year_Target", "Prior_Year_Actual",
    "Target_Value", "Variance_Band",
]

COL_WIDTHS = {
    "Practice_Name": 28, "Level": 14,
    "Site_ID": 12, "Site_Name": 28,
    "Practitioner_ID": 16, "Practitioner_Name": 28,
    "Metric_Key": 32, "Metric_Name": 38,
    "Section": 12, "Format_Type": 12,
    "Prior_Year_Target": 16, "Prior_Year_Actual": 16,
    "Target_Value": 14, "Variance_Band": 14,
}

INPUT_COLS = {"Target_Value", "Variance_Band"}


def instructions_text(practice_name, fy):
    return [
        (f"Dentally Analytics — Targets Template", True),
        (f"Practice: {practice_name}   |   Financial Year: {fy}", True),
        ("", False),
        ("HOW TO FILL IN THIS TEMPLATE", True),
        ("", False),
        ("1. Fill in Target_Value for the rows you want to set.", False),
        ("   Leave Target_Value blank for rows you do not want — only filled rows are loaded.", False),
        ("", False),
        ("2. HIERARCHY — targets are resolved in this order:", True),
        ("   Practice    = the default for all sites and practitioners.", False),
        ("   Site        = overrides Practice for that site only.", False),
        ("   Practitioner = overrides Site/Practice for that individual.", False),
        ("   You only need Site or Practitioner rows where the value differs.", False),
        ("", False),
        ("3. TARGET_VALUE FORMAT", True),
        ("   Format_Type = currency  →  enter the raw £ amount, e.g. 500000", False),
        ("   Format_Type = count     →  enter the raw count,    e.g. 1200", False),
        ("   Format_Type = percent   →  enter the percentage,   e.g. 85  (not 0.85)", False),
        ("", False),
        ("4. VARIANCE_BAND (optional — controls green/amber/red card colour)", True),
        ("   percent metrics  →  absolute pp,  e.g. 3  means ±3 percentage points", False),
        ("   currency/count   →  relative %,   e.g. 10 means ±10%", False),
        ("   Leave blank and the card will be white (no colour banding).", False),
        ("", False),
        ("5. REFERENCE COLUMNS (read-only — shown for guidance, NOT loaded):", True),
        ("   Prior_Year_Target  =  last year's target for this row.", False),
        ("   Prior_Year_Actual  =  last year's actual (from the analytics warehouse).", False),
        ("   Tip: generate with  --prefill rollover  (copies last year's target) or", False),
        ("        --prefill actuals  (copies last year's actual) into Target_Value,", False),
        ("        then adjust individual rows by hand.", False),
        ("", False),
        ("6. Do not rename, move or delete columns.", False),
        ("", False),
        ("7. Load with:  py Scripts/Load_Targets_From_Template.py --input <this file>", False),
    ]


# ── Excel writer ──────────────────────────────────────────────────────────────

def write_workbook(path, data_rows, practice_name, fy):
    wb = openpyxl.Workbook()

    # Instructions sheet
    ws_i = wb.active
    ws_i.title = "Instructions"
    ws_i.column_dimensions["A"].width = 80
    for i, (text, bold) in enumerate(instructions_text(practice_name, fy), 1):
        cell = ws_i.cell(row=i, column=1, value=text)
        cell.font = Font(bold=bold, size=11 if bold else 10)

    # Targets sheet
    ws = wb.create_sheet("Targets")

    for col_idx, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    input_font = Font(bold=True)
    dim_font   = Font(color="444444")
    for row_idx, row in enumerate(data_rows, 2):
        level = row[1]   # Level is index 1 in the row (after Practice_Name)
        lv_fill = PatternFill("solid", fgColor=LEVEL_FILLS[level])
        for col_idx, (h, val) in enumerate(zip(HEADERS, row), 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if h in INPUT_COLS:
                cell.font = input_font
            else:
                cell.fill = lv_fill
                cell.font = dim_font

    for col_idx, h in enumerate(HEADERS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[h]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    wb.save(path)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--server",   required=True)
    p.add_argument("--database", required=True)
    p.add_argument("--tenant",   required=True, type=int, help="Tenant_ID to generate for")
    p.add_argument("--username", default=None,
                   help="Azure AD login hint (used with --auth interactive)")
    p.add_argument("--auth",     default="default", choices=["default", "interactive"])
    p.add_argument("--fy",       required=True,
                   help="Financial year to generate, e.g. 2025-26")
    p.add_argument("--prefill",  default="same",
                   choices=["same", "rollover", "actuals", "blank"],
                   help="Default for Target_Value: same=this FY's existing target (default); "
                        "rollover=prior FY's target; actuals=prior FY's actual "
                        "(Gold.Fact_Metric_Actuals); blank=leave empty. The Prior_Year_Target "
                        "and Prior_Year_Actual reference columns are always shown so you can "
                        "choose per metric in Excel.")
    args = p.parse_args()

    # Output directory — Targets/ at repo root (parent of Scripts/)
    out_dir = Path(__file__).parent.parent / "Targets"
    out_dir.mkdir(exist_ok=True)

    if not re.match(r"^\d{4}-\d{2}$", args.fy):
        sys.exit("--fy must be in YYYY-YY format, e.g. 2025-26")
    years = [args.fy]
    tid = args.tenant

    print(f"Connecting to {args.server} / {args.database} ...")
    try:
        conn = connect(args.server, args.database, args.username, args.auth)
    except Exception as e:
        sys.exit(f"Connection failed: {e}")

    cur = conn.cursor()

    # Metrics
    cur.execute("""
        SELECT Metric_Key, Display_Name, Section, Format_Type,
               Supports_Site, Supports_Practitioner
        FROM Config.Metric_Definitions
        WHERE Is_Active = 1
        ORDER BY Display_Order
    """)
    metrics = cur.fetchall()

    # Practice name
    cur.execute("SELECT TOP 1 Practice_Name FROM Gold.Dim_Practice_Sites WHERE Tenant_ID = ?", [tid])
    row = cur.fetchone()
    if not row:
        sys.exit(f"Tenant {tid} not found in Gold.Dim_Practice_Sites.")
    practice_name = row.Practice_Name

    # Sites
    cur.execute("""
        SELECT Site_ID, Site_Name
        FROM Gold.Dim_Practice_Sites
        WHERE Tenant_ID = ? AND Site_Active = 1
        ORDER BY Site_Name
    """, [tid])
    sites = cur.fetchall()

    # Practitioners
    cur.execute("""
        SELECT Practitioner_ID, Full_Name, Site_ID
        FROM Gold.Dim_Practitioners
        WHERE Tenant_ID = ? AND Active = 1
        ORDER BY Full_Name
    """, [tid])
    practitioners = cur.fetchall()

    # Existing annual targets for pre-fill
    cur.execute("""
        SELECT Site_ID, Practitioner_ID, Metric, Period_Value, Target_Value, Variance
        FROM Input.Targets
        WHERE Tenant_ID = ? AND Period_Type = 'annual'
    """, [tid])
    existing = {
        (r.Site_ID, r.Practitioner_ID, r.Metric, _norm_fy(r.Period_Value)):
        (float(r.Target_Value), float(r.Variance) if r.Variance is not None else None)
        for r in cur.fetchall()
    }

    # Prior-year ACTUALS from Gold.Fact_Metric_Actuals. That fact is now DAILY grain
    # (fk_Date + Numerator/Denominator), not the old FY grain, so we aggregate per
    # (Metric, grain, FY) here, matching the DAX shapes:
    #   rate (any Denominator present)  -> SUM(Num)/SUM(Den)  (x100 for percent format)
    #   cumulative                      -> SUM(Num)
    #   point-in-time (no Denominator)  -> Num at the latest date in the FY
    # FY comes from Dim_Date.Financial_Year_Name; surrogate keys mapped to business keys
    # (-1 -> NULL = practice/all grain). Wrapped: the fact may not exist on a fresh install.
    actuals = {}
    try:
        cur.execute("""
            WITH base AS (
                SELECT a.Tenant_ID, a.fk_Practice_Site, a.fk_Practitioner, a.Metric,
                       d.Financial_Year_Name AS Period_Value, a.fk_Date,
                       a.Numerator, a.Denominator, md.Target_Type, md.Format_Type,
                       ROW_NUMBER() OVER (PARTITION BY a.Tenant_ID, a.fk_Practice_Site,
                            a.fk_Practitioner, a.Metric, d.Financial_Year_Name
                            ORDER BY a.fk_Date DESC) AS rn
                FROM Gold.Fact_Metric_Actuals a
                JOIN Gold.Dim_Date d            ON d.pk_Date    = a.fk_Date
                JOIN Config.Metric_Definitions md ON md.Metric_Key = a.Metric
                WHERE a.Tenant_ID = ?
            )
            SELECT
                CASE WHEN b.fk_Practice_Site = -1 THEN NULL ELSE ps.Site_ID END        AS Site_ID,
                CASE WHEN b.fk_Practitioner  = -1 THEN NULL ELSE pr.Practitioner_ID END AS Practitioner_ID,
                b.Metric, b.Period_Value,
                CASE
                    WHEN MAX(CASE WHEN b.Denominator IS NOT NULL THEN 1 ELSE 0 END) = 1
                        THEN (CASE WHEN b.Format_Type = 'percent' THEN 100.0 ELSE 1 END)
                             * SUM(b.Numerator) / NULLIF(SUM(b.Denominator), 0)
                    WHEN b.Target_Type = 'cumulative' THEN SUM(b.Numerator)
                    ELSE MAX(CASE WHEN b.rn = 1 THEN b.Numerator END)
                END AS Actual_Value
            FROM base b
            LEFT JOIN Gold.Dim_Practice_Sites ps ON ps.Tenant_ID = b.Tenant_ID AND ps.pk_Practice_Site = b.fk_Practice_Site
            LEFT JOIN Gold.Dim_Practitioners  pr ON pr.Tenant_ID = b.Tenant_ID AND pr.pk_Practitioner  = b.fk_Practitioner
            GROUP BY b.fk_Practice_Site, b.fk_Practitioner, ps.Site_ID, pr.Practitioner_ID,
                     b.Metric, b.Period_Value, b.Target_Type, b.Format_Type
        """, [tid])
        actuals = {
            (r.Site_ID, r.Practitioner_ID, r.Metric, _norm_fy(r.Period_Value)): float(r.Actual_Value)
            for r in cur.fetchall() if r.Actual_Value is not None
        }
    except Exception as e:
        print(f"  (Prior_Year_Actual unavailable -- Fact_Metric_Actuals not deployed? {e})")

    conn.close()

    site_name = {s.Site_ID: s.Site_Name for s in sites}

    # Generate one file per FY year
    for fy in years:
        period_value = "FY " + fy        # DB Period_Value 'FY 2025-26'
        prior_pv     = _prior_fy_pv(fy)  # 'FY 2024-25' -- rollover / actuals reference FY
        data_rows    = []

        def build_row(level, site_id, site_nm, prac_id, prac_nm, m):
            skey     = site_id or None   # '' (practice grain) -> None key
            pkey     = prac_id or None
            ev_same  = existing.get((skey, pkey, m.Metric_Key, period_value))
            ev_prior = existing.get((skey, pkey, m.Metric_Key, prior_pv))
            prior_actual = actuals.get((skey, pkey, m.Metric_Key, prior_pv))
            prior_target = ev_prior[0] if ev_prior else None
            # Default for the editable Target_Value (references shown regardless).
            if   args.prefill == "rollover":
                tv, vb = ev_prior if ev_prior else (None, None)
            elif args.prefill == "actuals":
                tv, vb = prior_actual, (ev_same[1] if ev_same else None)
            elif args.prefill == "blank":
                tv, vb = None, None
            else:  # 'same'
                tv, vb = ev_same if ev_same else (None, None)
            return [practice_name, level, site_id, site_nm, prac_id, prac_nm,
                    m.Metric_Key, m.Display_Name, m.Section, m.Format_Type,
                    prior_target, prior_actual, tv, vb]

        for m in metrics:
            data_rows.append(build_row("Practice", "", "", "", "", m))
            if m.Supports_Site:
                for s in sites:
                    data_rows.append(build_row("Site", s.Site_ID, s.Site_Name, "", "", m))
            if m.Supports_Practitioner:
                for pr in practitioners:
                    data_rows.append(build_row("Practitioner", "", "", pr.Practitioner_ID, pr.Full_Name, m))

        out_path = out_dir / f"T{tid}_{fy}.xlsx"
        write_workbook(out_path, data_rows, practice_name, fy)
        prefilled = sum(1 for r in data_rows if r[-2] is not None)
        print(f"Saved: {out_path}  ({len(data_rows):,} rows, {prefilled} pre-filled)")

    print("Done.")


if __name__ == "__main__":
    main()
